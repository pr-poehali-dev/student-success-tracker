#!/usr/bin/env python3
"""
Script to create an Excel file example for importing classes data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_classes_import_example():
    """Create an Excel file with example classes data."""
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Классы"
    
    # Define headers
    headers = [
        "ID класса",
        "Название класса",
        "Количество учеников",
        "Ответственный учитель ID",
        "Ответственный учитель ФИО"
    ]
    
    # Add headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Example data
    example_data = [
        ["class-001", "5А", 25, "teacher-001", "Иванов Иван Иванович"],
        ["class-002", "6Б", 28, "teacher-002", "Петрова Мария Сергеевна"],
        ["class-003", "7В", 23, "teacher-003", "Сидоров Алексей Петрович"],
        ["class-004", "8Г", 26, "teacher-004", "Козлова Елена Владимировна"],
    ]
    
    # Add example data
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left" if col_num != 3 else "center", vertical="center")
    
    # Adjust column widths
    column_widths = [18, 20, 25, 30, 35]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Ensure the public directory exists
    output_dir = "public"
    os.makedirs(output_dir, exist_ok=True)
    
    # Save the file
    output_path = os.path.join(output_dir, "import_example.xlsx")
    wb.save(output_path)
    
    print(f"✓ Excel file created successfully at: {output_path}")
    print(f"✓ Sheet name: {ws.title}")
    print(f"✓ Number of example rows: {len(example_data)}")
    print(f"✓ Columns: {', '.join(headers)}")

if __name__ == "__main__":
    create_classes_import_example()
