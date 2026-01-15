#!/usr/bin/env python3
"""
Create an Excel file example for importing classes data.
This script creates public/import_example.xlsx with sample data.

Requirements: openpyxl (pip install openpyxl)
Alternative: xlsxwriter (pip install xlsxwriter)

Run: python3 scripts/create_excel_import_example.py
"""

import sys
import os

def create_with_openpyxl():
    """Create Excel file using openpyxl library."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Классы"
        
        # Headers
        headers = ["ID класса", "Название класса", "Количество учеников", 
                   "Ответственный учитель ID", "Ответственный учитель ФИО"]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        data = [
            ["class-001", "5А", 25, "teacher-001", "Иванов Иван Иванович"],
            ["class-002", "6Б", 28, "teacher-002", "Петрова Мария Сергеевна"],
            ["class-003", "7В", 23, "teacher-003", "Сидоров Алексей Петрович"],
            ["class-004", "8Г", 26, "teacher-004", "Козлова Елена Владимировна"],
        ]
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left" if col_num != 3 else "center", 
                                         vertical="center")
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 23
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 35
        
        # Save
        os.makedirs("public", exist_ok=True)
        filepath = "public/import_example.xlsx"
        wb.save(filepath)
        
        print(f"✓ Excel file created successfully using openpyxl!")
        print(f"✓ Location: {os.path.abspath(filepath)}")
        print(f"✓ Sheet name: Классы")
        print(f"✓ Rows: {len(data)}")
        return True
        
    except ImportError:
        return False

def create_with_xlsxwriter():
    """Create Excel file using xlsxwriter library."""
    try:
        import xlsxwriter
        
        os.makedirs("public", exist_ok=True)
        filepath = "public/import_example.xlsx"
        
        workbook = xlsxwriter.Workbook(filepath)
        worksheet = workbook.add_worksheet("Классы")
        
        # Formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter'
        })
        
        text_format = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
        number_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        
        # Headers
        headers = ["ID класса", "Название класса", "Количество учеников",
                   "Ответственный учитель ID", "Ответственный учитель ФИО"]
        
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Data
        data = [
            ["class-001", "5А", 25, "teacher-001", "Иванов Иван Иванович"],
            ["class-002", "6Б", 28, "teacher-002", "Петрова Мария Сергеевна"],
            ["class-003", "7В", 23, "teacher-003", "Сидоров Алексей Петрович"],
            ["class-004", "8Г", 26, "teacher-004", "Козлова Елена Владимировна"],
        ]
        
        for row_num, row_data in enumerate(data, 1):
            for col_num, value in enumerate(row_data):
                fmt = number_format if col_num == 2 else text_format
                worksheet.write(row_num, col_num, value, fmt)
        
        # Column widths
        worksheet.set_column(0, 0, 15)
        worksheet.set_column(1, 1, 18)
        worksheet.set_column(2, 2, 23)
        worksheet.set_column(3, 3, 28)
        worksheet.set_column(4, 4, 35)
        
        workbook.close()
        
        print(f"✓ Excel file created successfully using xlsxwriter!")
        print(f"✓ Location: {os.path.abspath(filepath)}")
        print(f"✓ Sheet name: Классы")
        print(f"✓ Rows: {len(data)}")
        return True
        
    except ImportError:
        return False

def create_basic():
    """Create basic Excel file without external libraries (minimal format)."""
    import struct
    
    # This creates a minimal valid XLSX file
    # XLSX is actually a ZIP file with XML inside
    import zipfile
    from datetime import datetime
    
    try:
        os.makedirs("public", exist_ok=True)
        filepath = "public/import_example.xlsx"
        
        # Create basic XLSX structure
        with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as xlsx:
            # [Content_Types].xml
            content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>'''
            xlsx.writestr('[Content_Types].xml', content_types)
            
            # _rels/.rels
            rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''
            xlsx.writestr('_rels/.rels', rels)
            
            # xl/_rels/workbook.xml.rels
            wb_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>'''
            xlsx.writestr('xl/_rels/workbook.xml.rels', wb_rels)
            
            # xl/workbook.xml
            workbook = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets>
<sheet name="Классы" sheetId="1" r:id="rId1"/>
</sheets>
</workbook>'''
            xlsx.writestr('xl/workbook.xml', workbook)
            
            # xl/sharedStrings.xml
            shared_strings = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="20" uniqueCount="20">
<si><t>ID класса</t></si>
<si><t>Название класса</t></si>
<si><t>Количество учеников</t></si>
<si><t>Ответственный учитель ID</t></si>
<si><t>Ответственный учитель ФИО</t></si>
<si><t>class-001</t></si>
<si><t>5А</t></si>
<si><t>teacher-001</t></si>
<si><t>Иванов Иван Иванович</t></si>
<si><t>class-002</t></si>
<si><t>6Б</t></si>
<si><t>teacher-002</t></si>
<si><t>Петрова Мария Сергеевна</t></si>
<si><t>class-003</t></si>
<si><t>7В</t></si>
<si><t>teacher-003</t></si>
<si><t>Сидоров Алексей Петрович</t></si>
<si><t>class-004</t></si>
<si><t>8Г</t></si>
<si><t>teacher-004</t></si>
<si><t>Козлова Елена Владимировна</t></si>
</sst>'''
            xlsx.writestr('xl/sharedStrings.xml', shared_strings)
            
            # xl/worksheets/sheet1.xml
            worksheet = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetData>
<row r="1">
<c r="A1" t="s"><v>0</v></c>
<c r="B1" t="s"><v>1</v></c>
<c r="C1" t="s"><v>2</v></c>
<c r="D1" t="s"><v>3</v></c>
<c r="E1" t="s"><v>4</v></c>
</row>
<row r="2">
<c r="A2" t="s"><v>5</v></c>
<c r="B2" t="s"><v>6</v></c>
<c r="C2"><v>25</v></c>
<c r="D2" t="s"><v>7</v></c>
<c r="E2" t="s"><v>8</v></c>
</row>
<row r="3">
<c r="A3" t="s"><v>9</v></c>
<c r="B3" t="s"><v>10</v></c>
<c r="C3"><v>28</v></c>
<c r="D3" t="s"><v>11</v></c>
<c r="E3" t="s"><v>12</v></c>
</row>
<row r="4">
<c r="A4" t="s"><v>13</v></c>
<c r="B4" t="s"><v>14</v></c>
<c r="C4"><v>23</v></c>
<c r="D4" t="s"><v>15</v></c>
<c r="E4" t="s"><v>16</v></c>
</row>
<row r="5">
<c r="A5" t="s"><v>17</v></c>
<c r="B5" t="s"><v>18</v></c>
<c r="C5"><v>26</v></c>
<c r="D5" t="s"><v>19</v></c>
<c r="E5" t="s"><v>20</v></c>
</row>
</sheetData>
</worksheet>'''
            xlsx.writestr('xl/worksheets/sheet1.xml', worksheet)
        
        print(f"✓ Excel file created successfully using basic method!")
        print(f"✓ Location: {os.path.abspath(filepath)}")
        print(f"✓ Sheet name: Классы")
        print(f"✓ Rows: 4")
        return True
        
    except Exception as e:
        print(f"Error with basic method: {e}")
        return False

if __name__ == "__main__":
    print("Creating Excel import example...\n")
    
    if create_with_openpyxl():
        sys.exit(0)
    elif create_with_xlsxwriter():
        sys.exit(0)
    elif create_basic():
        sys.exit(0)
    else:
        print("\n✗ Failed to create Excel file.")
        print("\nPlease install one of these libraries:")
        print("  pip install openpyxl")
        print("  pip install xlsxwriter")
        sys.exit(1)
