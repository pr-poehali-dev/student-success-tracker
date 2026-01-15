#!/usr/bin/env python3
"""
Script to create an Excel file example for importing classes data.
This version tries openpyxl first, then falls back to xlsxwriter if needed.
"""

import os
import sys

def create_with_openpyxl():
    """Create using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Классы"
        
        headers = [
            "ID класса",
            "Название класса",
            "Количество учеников",
            "Ответственный учитель ID",
            "Ответственный учитель ФИО"
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        example_data = [
            ["class-001", "5А", 25, "teacher-001", "Иванов Иван Иванович"],
            ["class-002", "6Б", 28, "teacher-002", "Петрова Мария Сергеевна"],
            ["class-003", "7В", 23, "teacher-003", "Сидоров Алексей Петрович"],
            ["class-004", "8Г", 26, "teacher-004", "Козлова Елена Владимировна"],
        ]
        
        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left" if col_num != 3 else "center", vertical="center")
        
        column_widths = [18, 20, 25, 30, 35]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        output_dir = "public"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "import_example.xlsx")
        wb.save(output_path)
        
        print(f"SUCCESS: Excel file created at: {output_path}")
        print(f"Sheet name: {ws.title}")
        print(f"Number of example rows: {len(example_data)}")
        return True
        
    except ImportError:
        return False

def create_with_xlsxwriter():
    """Create using xlsxwriter."""
    try:
        import xlsxwriter
        
        output_dir = "public"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "import_example.xlsx")
        
        workbook = xlsxwriter.Workbook(output_path)
        worksheet = workbook.add_worksheet("Классы")
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter'
        })
        
        text_format = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
        number_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        
        headers = [
            "ID класса",
            "Название класса",
            "Количество учеников",
            "Ответственный учитель ID",
            "Ответственный учитель ФИО"
        ]
        
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        example_data = [
            ["class-001", "5А", 25, "teacher-001", "Иванов Иван Иванович"],
            ["class-002", "6Б", 28, "teacher-002", "Петрова Мария Сергеевна"],
            ["class-003", "7В", 23, "teacher-003", "Сидоров Алексей Петрович"],
            ["class-004", "8Г", 26, "teacher-004", "Козлова Елена Владимировна"],
        ]
        
        for row_num, row_data in enumerate(example_data, 1):
            for col_num, value in enumerate(row_data):
                fmt = number_format if col_num == 2 else text_format
                worksheet.write(row_num, col_num, value, fmt)
        
        worksheet.set_column(0, 0, 18)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 25)
        worksheet.set_column(3, 3, 30)
        worksheet.set_column(4, 4, 35)
        
        workbook.close()
        
        print(f"SUCCESS: Excel file created at: {output_path}")
        print(f"Sheet name: Классы")
        print(f"Number of example rows: {len(example_data)}")
        return True
        
    except ImportError:
        return False

if __name__ == "__main__":
    if create_with_openpyxl():
        sys.exit(0)
    elif create_with_xlsxwriter():
        sys.exit(0)
    else:
        print("ERROR: Neither openpyxl nor xlsxwriter is installed.")
        print("Please install one of them:")
        print("  pip install openpyxl")
        print("  pip install xlsxwriter")
        sys.exit(1)
